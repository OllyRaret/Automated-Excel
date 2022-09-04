import pandas as pd
import string
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

def excel_automated(file_name):
    # !! file should be named after the following template: sales_period.xlsx !!
    excel_file = pd.read_excel(file_name)
    # getting time period
    period_and_extension = file_name.split('_')[1]
    period = period_and_extension.split('.')[0]

    report_table = excel_file.pivot_table(index='Gender', columns='Product line',
                                          values='Total', aggfunc='sum').round(0)
    report_table.to_excel(f'report_{period_and_extension}', sheet_name='Report', startrow=4)
    wb = load_workbook(f'report_{period_and_extension}')
    sheet = wb['Report']

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # giving titles
    sheet['A1'] = 'Sales Report'
    sheet['A1'].font = Font('Castellar', bold=True, size=20)
    sheet['A2'] = period.title()
    sheet['A2'].font = Font('Castellar', bold=True, size=10)

    # creating a barchart
    chart = BarChart()
    # locate categories and data
    categories = Reference(sheet, min_col=min_column, max_col=min_column,
                           min_row=min_row + 1, max_row=max_row)  # not including headers
    data = Reference(sheet, min_col=min_column + 1, max_col=max_column,
                     min_row=min_row, max_row=max_row)  # including headers
    # adding data and categories
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    # location chart
    sheet.add_chart(chart, "B12")
    chart.title = 'Sales by Product'
    chart.style = 15  # choose the chart style

    # applying result formulas
    # first create alphabet list as references for cells
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column]
    # summary in columns B-G
    for i in excel_alphabet:
        if i != 'A':
            sheet[f'{i}{max_row + 1}'] = f'=SUM({i}{min_row + 1}:{i}{max_row})'
            sheet[f'{i}{max_row + 1}'].style = 'Currency'
    # adding total label
    sheet[f'{excel_alphabet[0]}{max_row + 1}'] = 'Total'

    # saving file
    wb.save(f'report_{period_and_extension}')
    return

excel_automated('sales_december.xlsx')